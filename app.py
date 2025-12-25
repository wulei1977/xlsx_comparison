"""
Excel文件对比Web服务
"""

import os
import sys
import uuid
import webbrowser
import threading
import zipfile
import re
from datetime import datetime
from flask import Flask, render_template, request, jsonify, send_file
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font
from openpyxl.comments import Comment


def fix_xlsx_for_openpyxl(input_path: str) -> str:
    """
    修复xlsx文件中dataValidations元素的非标准属性，使openpyxl能正常加载。
    某些自动化工具生成的Excel文件会在dataValidations元素上添加algorithmName等属性，
    这些属性不被openpyxl支持，需要移除。
    返回修复后的临时文件路径。
    """
    temp_path = input_path + '.fixed.xlsx'
    with zipfile.ZipFile(input_path, 'r') as zin:
        with zipfile.ZipFile(temp_path, 'w', zipfile.ZIP_DEFLATED) as zout:
            for item in zin.namelist():
                content = zin.read(item)
                if item.endswith('.xml') and b'dataValidations' in content:
                    content_str = content.decode('utf-8')
                    # 移除 dataValidations 元素中的非标准属性
                    content_str = re.sub(r'(<dataValidations[^>]*?)\s+algorithmName="[^"]*"', r'\1', content_str)
                    content_str = re.sub(r'(<dataValidations[^>]*?)\s+hashValue="[^"]*"', r'\1', content_str)
                    content_str = re.sub(r'(<dataValidations[^>]*?)\s+saltValue="[^"]*"', r'\1', content_str)
                    content_str = re.sub(r'(<dataValidations[^>]*?)\s+spinCount="[^"]*"', r'\1', content_str)
                    content = content_str.encode('utf-8')
                zout.writestr(item, content)
    return temp_path


def get_base_path():
    """获取基础路径，支持PyInstaller打包"""
    if getattr(sys, 'frozen', False):
        return sys._MEIPASS
    return os.path.dirname(os.path.abspath(__file__))


def get_upload_folder():
    """获取上传文件夹路径"""
    if getattr(sys, 'frozen', False):
        # 打包后使用exe所在目录
        base = os.path.dirname(sys.executable)
    else:
        base = os.path.dirname(os.path.abspath(__file__))
    return os.path.join(base, 'uploads')


base_path = get_base_path()

# 从命令行参数或环境变量获取子路径前缀
import argparse
parser = argparse.ArgumentParser()
parser.add_argument('--prefix', '-p', default='', help='URL路径前缀，如 excel-compare')
args, _ = parser.parse_known_args()

SCRIPT_NAME = args.prefix or os.environ.get('SCRIPT_NAME', 'excel-compare')
SCRIPT_NAME = SCRIPT_NAME.strip('/')
if SCRIPT_NAME:
    SCRIPT_NAME = '/' + SCRIPT_NAME

app = Flask(__name__, template_folder=os.path.join(base_path, 'templates'))
app.config['UPLOAD_FOLDER'] = get_upload_folder()
app.config['MAX_CONTENT_LENGTH'] = 50 * 1024 * 1024  # 50MB

os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)


@app.context_processor
def inject_base_path():
    """注入基础路径到模板"""
    return {'base_path': SCRIPT_NAME}


def load_excel(file_path: str, sheet_name: str) -> pd.DataFrame:
    """加载Excel文件的指定worksheet"""
    return pd.read_excel(file_path, sheet_name=sheet_name, engine="calamine")


def get_excel_info(file_path: str) -> dict:
    """获取Excel文件的sheet名和列名"""
    xl = pd.ExcelFile(file_path, engine="calamine")
    sheets_info = {}
    for sheet in xl.sheet_names:
        df = pd.read_excel(xl, sheet_name=sheet, nrows=0, engine="calamine")
        sheets_info[sheet] = list(df.columns)
    return sheets_info


def create_composite_key(df: pd.DataFrame, key_columns: list) -> pd.Series:
    """根据多个列创建组合键"""
    return df[key_columns].astype(str).agg("||".join, axis=1)


def mark_excel_differences(
    file1_path: str, file2_path: str,
    sheet1: str, sheet2: str,
    key_columns: list,
    output1_path: str, output2_path: str
) -> dict:
    """标注两个Excel文件的差异并保存"""
    from copy import copy
    
    # 加载数据用于对比
    df1 = load_excel(file1_path, sheet1)
    df2 = load_excel(file2_path, sheet2)
    
    # 保存原始列（不包含辅助列）
    original_cols1 = list(df1.columns)
    original_cols2 = list(df2.columns)
    
    df1["_composite_key"] = create_composite_key(df1, key_columns)
    df2["_composite_key"] = create_composite_key(df2, key_columns)
    df1["_row_num"] = df1.index + 2
    df2["_row_num"] = df2.index + 2
    
    # 创建键到行号的映射（处理重复键，只取第一个）
    key_to_row1 = {}
    key_to_row2 = {}
    for _, row in df1.iterrows():
        key = row["_composite_key"]
        if key not in key_to_row1:
            key_to_row1[key] = int(row["_row_num"])
    for _, row in df2.iterrows():
        key = row["_composite_key"]
        if key not in key_to_row2:
            key_to_row2[key] = int(row["_row_num"])
    
    keys1 = set(df1["_composite_key"])
    keys2 = set(df2["_composite_key"])
    
    only_in_file1 = keys1 - keys2
    only_in_file2 = keys2 - keys1
    common_keys = keys1 & keys2
    
    # 定义样式
    green_fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    red_font = Font(color="FF0000")
    
    # 获取列名到列索引的映射
    col_to_idx1 = {col: idx + 1 for idx, col in enumerate(original_cols1)}
    col_to_idx2 = {col: idx + 1 for idx, col in enumerate(original_cols2)}
    
    # 获取key列的索引
    key_col_indices1 = [col_to_idx1[k] for k in key_columns if k in col_to_idx1]
    key_col_indices2 = [col_to_idx2[k] for k in key_columns if k in col_to_idx2]
    
    # 加载原始Excel文件（用openpyxl保留样式）
    # 先修复可能存在的非标准属性问题
    fixed_file1 = fix_xlsx_for_openpyxl(file1_path)
    fixed_file2 = fix_xlsx_for_openpyxl(file2_path)
    
    try:
        wb1_orig = load_workbook(fixed_file1)
        wb2_orig = load_workbook(fixed_file2)
    finally:
        # 清理临时文件
        if os.path.exists(fixed_file1):
            os.remove(fixed_file1)
        if os.path.exists(fixed_file2):
            os.remove(fixed_file2)
    
    ws1_orig = wb1_orig[sheet1]
    ws2_orig = wb2_orig[sheet2]
    
    # 创建新工作簿，复制指定sheet的内容和样式
    from openpyxl import Workbook
    wb1 = Workbook()
    ws1 = wb1.active
    ws1.title = sheet1
    
    wb2 = Workbook()
    ws2 = wb2.active
    ws2.title = sheet2
    
    # 复制sheet1的内容、样式、行高、列宽
    for row_idx, row in enumerate(ws1_orig.iter_rows(), 1):
        for col_idx, cell in enumerate(row, 1):
            new_cell = ws1.cell(row=row_idx, column=col_idx, value=cell.value)
            if cell.has_style:
                new_cell.font = copy(cell.font)
                new_cell.fill = copy(cell.fill)
                new_cell.border = copy(cell.border)
                new_cell.alignment = copy(cell.alignment)
                new_cell.number_format = cell.number_format
                new_cell.protection = copy(cell.protection)
    
    # 复制列宽
    for col_letter, col_dim in ws1_orig.column_dimensions.items():
        ws1.column_dimensions[col_letter].width = col_dim.width
    # 复制行高
    for row_idx, row_dim in ws1_orig.row_dimensions.items():
        ws1.row_dimensions[row_idx].height = row_dim.height
    
    # 复制sheet2的内容、样式、行高、列宽
    for row_idx, row in enumerate(ws2_orig.iter_rows(), 1):
        for col_idx, cell in enumerate(row, 1):
            new_cell = ws2.cell(row=row_idx, column=col_idx, value=cell.value)
            if cell.has_style:
                new_cell.font = copy(cell.font)
                new_cell.fill = copy(cell.fill)
                new_cell.border = copy(cell.border)
                new_cell.alignment = copy(cell.alignment)
                new_cell.number_format = cell.number_format
                new_cell.protection = copy(cell.protection)
    
    # 复制列宽
    for col_letter, col_dim in ws2_orig.column_dimensions.items():
        ws2.column_dimensions[col_letter].width = col_dim.width
    # 复制行高
    for row_idx, row_dim in ws2_orig.row_dimensions.items():
        ws2.row_dimensions[row_idx].height = row_dim.height
    
    # 关闭原始工作簿
    wb1_orig.close()
    wb2_orig.close()
    
    # 标注仅在文件1中存在的行
    for key in only_in_file1:
        row_num = key_to_row1[key]
        for col_idx in range(1, len(original_cols1) + 1):
            cell = ws1.cell(row=row_num, column=col_idx)
            cell.fill = green_fill
        if key_col_indices1:
            cell = ws1.cell(row=row_num, column=key_col_indices1[0])
            cell.comment = Comment("仅在此文件中存在的行", "Excel对比工具")
    
    # 标注仅在文件2中存在的行
    for key in only_in_file2:
        row_num = key_to_row2[key]
        for col_idx in range(1, len(original_cols2) + 1):
            cell = ws2.cell(row=row_num, column=col_idx)
            cell.fill = green_fill
        if key_col_indices2:
            cell = ws2.cell(row=row_num, column=key_col_indices2[0])
            cell.comment = Comment("仅在此文件中存在的行", "Excel对比工具")
    
    # 标注共有行的数据差异
    common_columns = [c for c in original_cols1 if c in original_cols2]
    
    # 为每个键获取第一行数据
    df1_first = df1.drop_duplicates(subset=["_composite_key"], keep="first").set_index("_composite_key")
    df2_first = df2.drop_duplicates(subset=["_composite_key"], keep="first").set_index("_composite_key")
    
    diff_row_count = 0
    for key in common_keys:
        row1 = df1_first.loc[key]
        row2 = df2_first.loc[key]
        row_num1 = int(row1["_row_num"])
        row_num2 = int(row2["_row_num"])
        
        has_diff = False
        for col in common_columns:
            val1 = row1[col] if col in row1.index else None
            val2 = row2[col] if col in row2.index else None
            
            val1_is_nan = pd.isna(val1)
            val2_is_nan = pd.isna(val2)
            
            if val1_is_nan and val2_is_nan:
                continue
            elif val1_is_nan != val2_is_nan or val1 != val2:
                has_diff = True
                # 标注文件1中的差异单元格
                if col in col_to_idx1:
                    cell1 = ws1.cell(row=row_num1, column=col_to_idx1[col])
                    cell1.fill = yellow_fill
                    cell1.font = red_font
                    comment_text = f"与文件2第{row_num2}行[{col}]列不同\n文件2值: {val2}"
                    cell1.comment = Comment(comment_text, "Excel对比工具")
                
                # 标注文件2中的差异单元格
                if col in col_to_idx2:
                    cell2 = ws2.cell(row=row_num2, column=col_to_idx2[col])
                    cell2.fill = yellow_fill
                    cell2.font = red_font
                    comment_text = f"与文件1第{row_num1}行[{col}]列不同\n文件1值: {val1}"
                    cell2.comment = Comment(comment_text, "Excel对比工具")
        
        if has_diff:
            diff_row_count += 1
    
    # 保存
    wb1.save(output1_path)
    wb2.save(output2_path)
    
    return {
        "only_in_file1": len(only_in_file1),
        "only_in_file2": len(only_in_file2),
        "common_with_diff": diff_row_count
    }


def compare_xlsx(file1: str, file2: str, sheet1: str, sheet2: str, key_columns: list) -> str:
    """对比两个Excel文件，返回对比结果文本"""
    lines = []
    
    lines.append("=" * 60)
    lines.append("Excel文件对比结果")
    lines.append(f"对比时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    lines.append("=" * 60)
    lines.append(f"文件1 Sheet: {sheet1}")
    lines.append(f"文件2 Sheet: {sheet2}")
    lines.append(f"组合键列: {key_columns}")
    lines.append("-" * 60)
    
    try:
        df1 = load_excel(file1, sheet1)
        df2 = load_excel(file2, sheet2)
    except Exception as e:
        return f"加载文件失败: {e}"
    
    lines.append(f"文件1行数: {len(df1)}, 列数: {len(df1.columns)}")
    lines.append(f"文件2行数: {len(df2)}, 列数: {len(df2.columns)}")
    
    # 验证键列存在
    for col in key_columns:
        if col not in df1.columns:
            return f"文件1中不存在列: {col}"
        if col not in df2.columns:
            return f"文件2中不存在列: {col}"
    
    # 创建组合键，保留原始行号（Excel行号从2开始，因为第1行是表头）
    df1["_composite_key"] = create_composite_key(df1, key_columns)
    df2["_composite_key"] = create_composite_key(df2, key_columns)
    df1["_row_num"] = df1.index + 2  # Excel行号
    df2["_row_num"] = df2.index + 2
    
    # 创建键到行号的映射
    key_to_row1 = dict(zip(df1["_composite_key"], df1["_row_num"]))
    key_to_row2 = dict(zip(df2["_composite_key"], df2["_row_num"]))
    
    keys1 = set(df1["_composite_key"])
    keys2 = set(df2["_composite_key"])
    
    only_in_file1 = keys1 - keys2
    only_in_file2 = keys2 - keys1
    common_keys = keys1 & keys2
    
    lines.append("-" * 60)
    lines.append("行级别差异统计:")
    lines.append(f"  仅在文件1中存在的行: {len(only_in_file1)}")
    lines.append(f"  仅在文件2中存在的行: {len(only_in_file2)}")
    lines.append(f"  两文件共有的行: {len(common_keys)}")
    
    if only_in_file1:
        lines.append("-" * 60)
        lines.append("仅在文件1中存在的行:")
        for key in sorted(only_in_file1):
            row_num = key_to_row1.get(key, "?")
            lines.append(f"  [文件1第{row_num}行] 键值: {key}")
    
    if only_in_file2:
        lines.append("-" * 60)
        lines.append("仅在文件2中存在的行:")
        for key in sorted(only_in_file2):
            row_num = key_to_row2.get(key, "?")
            lines.append(f"  [文件2第{row_num}行] 键值: {key}")
    
    lines.append("-" * 60)
    lines.append("共有行的数据差异:")
    
    common_columns = [c for c in df1.columns if c in df2.columns and c not in ("_composite_key", "_row_num")]
    diff_count = 0
    
    df1_indexed = df1.set_index("_composite_key")
    df2_indexed = df2.set_index("_composite_key")
    
    for key in sorted(common_keys):
        row1 = df1_indexed.loc[key]
        row2 = df2_indexed.loc[key]
        row_num1 = int(row1["_row_num"]) if "_row_num" in row1.index else "?"
        row_num2 = int(row2["_row_num"]) if "_row_num" in row2.index else "?"
        
        row_diffs = []
        for col in common_columns:
            val1 = row1[col] if col in row1.index else None
            val2 = row2[col] if col in row2.index else None
            
            val1_is_nan = pd.isna(val1)
            val2_is_nan = pd.isna(val2)
            
            if val1_is_nan and val2_is_nan:
                continue
            elif val1_is_nan != val2_is_nan or val1 != val2:
                row_diffs.append((col, val1, val2))
        
        if row_diffs:
            diff_count += 1
            lines.append(f"  键值: {key} [文件1第{row_num1}行 vs 文件2第{row_num2}行]")
            for col, v1, v2 in row_diffs:
                lines.append(f"    列[{col}]: 文件1='{v1}' vs 文件2='{v2}'")
    
    if diff_count == 0:
        lines.append("  无数据差异")
    
    # 列差异
    cols1 = set(df1.columns) - {"_composite_key", "_row_num"}
    cols2 = set(df2.columns) - {"_composite_key", "_row_num"}
    only_cols1 = cols1 - cols2
    only_cols2 = cols2 - cols1
    
    if only_cols1 or only_cols2:
        lines.append("-" * 60)
        lines.append("列级别差异:")
        if only_cols1:
            lines.append(f"  仅在文件1中存在的列: {sorted(only_cols1)}")
        if only_cols2:
            lines.append(f"  仅在文件2中存在的列: {sorted(only_cols2)}")
    
    lines.append("=" * 60)
    lines.append("对比完成")
    lines.append("=" * 60)
    
    return "\n".join(lines)


@app.route('/')
def index():
    return render_template('index.html', base_path=SCRIPT_NAME)


# 存储原始文件名的映射
file_original_names = {}

@app.route('/upload', methods=['POST'])
def upload_file():
    """上传文件并返回sheet和列信息"""
    if 'file' not in request.files:
        return jsonify({'error': '没有文件'}), 400
    
    file = request.files['file']
    if file.filename == '':
        return jsonify({'error': '没有选择文件'}), 400
    
    # 保存文件
    file_id = str(uuid.uuid4())
    filename = f"{file_id}.xlsx"
    filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
    file.save(filepath)
    
    # 保存原始文件名
    original_name = file.filename
    file_original_names[file_id] = original_name
    
    try:
        info = get_excel_info(filepath)
        return jsonify({
            'file_id': file_id,
            'original_name': original_name,
            'sheets': info
        })
    except Exception as e:
        os.remove(filepath)
        return jsonify({'error': str(e)}), 400


@app.route('/compare', methods=['POST'])
def compare():
    """执行对比"""
    data = request.json
    
    file1_id = data.get('file1_id')
    file2_id = data.get('file2_id')
    sheet1 = data.get('sheet1')
    sheet2 = data.get('sheet2')
    keys = data.get('keys', [])
    
    if not all([file1_id, file2_id, sheet1, sheet2, keys]):
        return jsonify({'error': '参数不完整'}), 400
    
    file1_path = os.path.join(app.config['UPLOAD_FOLDER'], f"{file1_id}.xlsx")
    file2_path = os.path.join(app.config['UPLOAD_FOLDER'], f"{file2_id}.xlsx")
    
    if not os.path.exists(file1_path) or not os.path.exists(file2_path):
        return jsonify({'error': '文件不存在'}), 400
    
    result = compare_xlsx(file1_path, file2_path, sheet1, sheet2, keys)
    
    # 保存结果文件
    result_id = str(uuid.uuid4())
    result_path = os.path.join(app.config['UPLOAD_FOLDER'], f"result_{result_id}.txt")
    with open(result_path, 'w', encoding='utf-8') as f:
        f.write(result)
    
    # 获取原始文件名（去掉扩展名）
    orig_name1 = file_original_names.get(file1_id, "文件1")
    orig_name2 = file_original_names.get(file2_id, "文件2")
    if orig_name1.lower().endswith('.xlsx'):
        orig_name1 = orig_name1[:-5]
    elif orig_name1.lower().endswith('.xls'):
        orig_name1 = orig_name1[:-4]
    if orig_name2.lower().endswith('.xlsx'):
        orig_name2 = orig_name2[:-5]
    elif orig_name2.lower().endswith('.xls'):
        orig_name2 = orig_name2[:-4]
    
    # 生成标注差异的Excel文件
    marked_file1_path = os.path.join(app.config['UPLOAD_FOLDER'], f"marked1_{result_id}.xlsx")
    marked_file2_path = os.path.join(app.config['UPLOAD_FOLDER'], f"marked2_{result_id}.xlsx")
    
    # 保存标注文件的下载名
    marked_names = {
        result_id: {
            1: f"{orig_name1}（标注）.xlsx",
            2: f"{orig_name2}（标注）.xlsx"
        }
    }
    # 存储到全局变量
    if not hasattr(app, 'marked_file_names'):
        app.marked_file_names = {}
    app.marked_file_names[result_id] = marked_names[result_id]
    
    try:
        mark_excel_differences(
            file1_path, file2_path,
            sheet1, sheet2,
            keys,
            marked_file1_path, marked_file2_path
        )
        has_marked_files = True
    except Exception as e:
        has_marked_files = False
        print(f"生成标注文件失败: {e}")
    
    return jsonify({
        'result': result,
        'result_id': result_id,
        'has_marked_files': has_marked_files
    })


@app.route('/download/<result_id>')
def download(result_id):
    """下载对比结果"""
    result_path = os.path.join(app.config['UPLOAD_FOLDER'], f"result_{result_id}.txt")
    if not os.path.exists(result_path):
        return "文件不存在", 404
    
    return send_file(
        result_path,
        as_attachment=True,
        download_name=f"compare_result_{datetime.now().strftime('%Y%m%d_%H%M%S')}.txt"
    )


@app.route('/download_marked/<result_id>/<int:file_num>')
def download_marked(result_id, file_num):
    """下载标注差异的Excel文件"""
    if file_num not in (1, 2):
        return "无效的文件编号", 400
    
    marked_path = os.path.join(app.config['UPLOAD_FOLDER'], f"marked{file_num}_{result_id}.xlsx")
    if not os.path.exists(marked_path):
        return "文件不存在", 404
    
    # 获取保存的下载文件名
    download_name = f"marked_file{file_num}.xlsx"
    if hasattr(app, 'marked_file_names') and result_id in app.marked_file_names:
        download_name = app.marked_file_names[result_id].get(file_num, download_name)
    
    return send_file(
        marked_path,
        as_attachment=True,
        download_name=download_name
    )


if __name__ == '__main__':
    port = 5000
    
    # 打包后自动打开浏览器
    if getattr(sys, 'frozen', False):
        if SCRIPT_NAME:
            url = f'http://127.0.0.1:{port}{SCRIPT_NAME}'
        else:
            url = f'http://127.0.0.1:{port}'
        threading.Timer(1.5, lambda: webbrowser.open(url)).start()
        print(f"Excel对比工具已启动，浏览器将自动打开: {url}")
        if SCRIPT_NAME:
            print(f"URL前缀: {SCRIPT_NAME}")
        print("关闭此窗口将停止服务")
        
        # 如果有子路径，使用 DispatcherMiddleware
        if SCRIPT_NAME:
            from werkzeug.middleware.dispatcher import DispatcherMiddleware
            from werkzeug.serving import run_simple
            
            def not_found(environ, start_response):
                start_response('404 Not Found', [('Content-Type', 'text/plain')])
                return [b'Not Found']
            
            application = DispatcherMiddleware(not_found, {SCRIPT_NAME: app})
            run_simple('0.0.0.0', port, application, use_reloader=False)
        else:
            app.run(host='0.0.0.0', port=port, debug=False)
    else:
        if SCRIPT_NAME:
            url = f'http://127.0.0.1:{port}{SCRIPT_NAME}'
            from werkzeug.middleware.dispatcher import DispatcherMiddleware
            from werkzeug.serving import run_simple
            
            def not_found(environ, start_response):
                start_response('404 Not Found', [('Content-Type', 'text/plain')])
                return [b'Not Found']
            
            application = DispatcherMiddleware(not_found, {SCRIPT_NAME: app})
            run_simple('0.0.0.0', port, application, use_reloader=True, use_debugger=True)
        else:
            app.run(debug=True, port=port)
